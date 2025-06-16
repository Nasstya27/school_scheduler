import openpyxl

def export_to_excel(schedule, filename="schedule.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница"]

    sheet['A1'] = "Класс"
    for i in range(len(days)):
        sheet.cell(row=1, column=i * 6 + 2).value = days[i]
        for j in range(1, 7):
            sheet.cell(row=1, column=i * 6 + 1 + j + 1).value = str(j)

    row_num = 2
    for class_name, daily_schedule in schedule.items():
        sheet.cell(row=row_num, column=1).value = class_name
        col_offset = 1
        for day in days:
            for lesson_num, lesson in enumerate(daily_schedule.get(day, [])):
                if lesson:
                    sheet.cell(row=row_num, column=col_offset + lesson_num + 2).value = \
                        f"{lesson['subject']} ({lesson['teacher']})"
                else:
                    sheet.cell(row=row_num, column=col_offset + lesson_num + 2).value = ""  

            col_offset += 6  
        row_num += 1

    workbook.save(filename)
    print(f"Расписание сохранено в {filename}")